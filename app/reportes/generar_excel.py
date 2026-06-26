"""Generación de reportes Excel con openpyxl (CU-28)."""

import io
from datetime import datetime, timedelta, timezone

_TZ = timezone(timedelta(hours=-5))  # America/Guayaquil — sin DST

import openpyxl
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from app.shared.analytics import load_agg, load_enriched_fact

_MESES_LABEL = [
    "",
    "Enero",
    "Febrero",
    "Marzo",
    "Abril",
    "Mayo",
    "Junio",
    "Julio",
    "Agosto",
    "Septiembre",
    "Octubre",
    "Noviembre",
    "Diciembre",
]
_MESES_SHORT = ["", "Ene", "Feb", "Mar", "Abr", "May", "Jun", "Jul", "Ago", "Sep", "Oct", "Nov", "Dic"]
_DOW_LABEL = ["", "Lunes", "Martes", "Miércoles", "Jueves", "Viernes", "Sábado", "Domingo"]
_FAA_DESC = {
    "A": "Aerolínea (Carrier)",
    "B": "Meteorología (Weather)",
    "C": "Sistema Aéreo (NAS)",
    "D": "Seguridad (Security)",
}
_QUARTERS = {"1": "Q1 (Ene–Mar)", "2": "Q2 (Abr–Jun)", "3": "Q3 (Jul–Sep)", "4": "Q4 (Oct–Dic)"}
_CAUSA_COLS = [
    ("carrierdelay", "Carrier (Aerolínea)"),
    ("weatherdelay", "Weather (Clima)"),
    ("nasdelay", "NAS (Sist. Aéreo)"),
    ("securitydelay", "Security (Seguridad)"),
    ("lateaircraftdelay", "LateAircraft (Avión tardío)"),
]

_BLUE_DARK = "1B3A6B"
_BLUE_LIGHT = "EBF0FA"
_GREEN = "16a34a"
_AMBER = "d97706"
_RED = "dc2626"
_THIN = Border(
    left=Side(style="thin", color="D1D5DB"),
    right=Side(style="thin", color="D1D5DB"),
    top=Side(style="thin", color="D1D5DB"),
    bottom=Side(style="thin", color="D1D5DB"),
)


# ── Style helpers ─────────────────────────────────────────────────────────────


def _hdr(ws, fila: int, columnas: list[str]) -> None:
    fill = PatternFill("solid", fgColor=_BLUE_DARK)
    font = Font(bold=True, color="FFFFFF", size=10)
    for i, col in enumerate(columnas, 1):
        cell = ws.cell(row=fila, column=i, value=col)
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center")
        cell.border = _THIN


def _autofit(ws) -> None:
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                max_len = max(max_len, len(str(cell.value or "")))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max_len + 4, 44)


def _color_otp(ws, row: int, col: int, val: float) -> None:
    color = _GREEN if val >= 85 else _AMBER if val >= 70 else _RED
    cell = ws.cell(row=row, column=col)
    cell.fill = PatternFill("solid", fgColor=color)
    cell.font = Font(bold=True, color="FFFFFF", size=10)
    cell.alignment = Alignment(horizontal="center")


def _stripe(ws, row: int, n_cols: int) -> None:
    if row % 2 == 0:
        fill = PatternFill("solid", fgColor=_BLUE_LIGHT)
        for c in range(1, n_cols + 1):
            ws.cell(row=row, column=c).fill = fill


def _filtros_display(filtros: dict | None) -> list[tuple[str, str]]:
    if not filtros:
        return [("Período", "Todos los datos")]
    pairs = []
    if filtros.get("year"):
        pairs.append(("Año", str(filtros["year"])))
    if filtros.get("quarter"):
        pairs.append(("Trimestre", _QUARTERS.get(str(filtros["quarter"]), filtros["quarter"])))
    if filtros.get("month"):
        m = int(filtros["month"])
        pairs.append(("Mes", _MESES_LABEL[m] if 1 <= m <= 12 else str(m)))
    if filtros.get("dow"):
        d = int(filtros["dow"])
        pairs.append(("Día de semana", _DOW_LABEL[d] if 1 <= d <= 7 else str(d)))
    if filtros.get("airline"):
        pairs.append(("Aerolínea", filtros["airline"]))
    if filtros.get("origin"):
        pairs.append(("Aeropuerto origen", filtros["origin"]))
    if filtros.get("dest"):
        pairs.append(("Aeropuerto destino", filtros["dest"]))
    if filtros.get("cancel_code"):
        pairs.append(("Causa cancelación", _FAA_DESC.get(filtros["cancel_code"], filtros["cancel_code"])))
    if filtros.get("solo_cancelados"):
        pairs.append(("Solo cancelados", "Sí"))
    return pairs or [("Período", "Todos los datos")]


# ── Sheet builders ────────────────────────────────────────────────────────────


def _sheet_otp(wb, df) -> None:
    """Hoja: Puntualidad OTP por aerolínea + bar chart."""
    ws = wb.create_sheet("Puntualidad OTP")
    if "Reporting_Airline" not in df.columns or "ArrDel15" not in df.columns:
        ws["A1"] = "Datos no disponibles"
        return

    vuelos_op = df[df["Cancelled"] == 0] if "Cancelled" in df.columns else df
    grp = (
        vuelos_op.groupby("Reporting_Airline")
        .agg(
            total=("pk_vuelo", "count"),
            otp_sum=("ArrDel15", lambda x: (x == 0).sum()),
        )
        .reset_index()
    )
    if "ArrDelayMinutes" in vuelos_op.columns:
        delay = (
            vuelos_op.groupby("Reporting_Airline")["ArrDelayMinutes"].mean().round(2).reset_index(name="retraso_prom")
        )
        grp = grp.merge(delay, on="Reporting_Airline", how="left")
    else:
        grp["retraso_prom"] = 0.0

    if "DepDel15" in vuelos_op.columns:
        dep = (
            vuelos_op.groupby("Reporting_Airline")
            .agg(
                dep_sum=("DepDel15", lambda x: (x == 1).sum()),
                dep_total=("pk_vuelo", "count"),
            )
            .reset_index()
        )
        dep["dep_pct"] = (dep["dep_sum"] / dep["dep_total"].replace(0, 1) * 100).round(2)
        grp = grp.merge(dep[["Reporting_Airline", "dep_pct"]], on="Reporting_Airline", how="left")

    grp["otp_pct"] = (grp["otp_sum"] / grp["total"] * 100).round(2)
    grp = grp.sort_values("total", ascending=False)

    has_dep = "dep_pct" in grp.columns
    headers = ["Aerolínea", "Vuelos op.", "OTP arr. (%)", "Retraso arr. prom. (min)"]
    if has_dep:
        headers.append("Dep. delay (%)")
    _hdr(ws, 1, headers)
    OTP_COL = 3

    for _, r in grp.iterrows():
        otp_val = round(float(r["otp_pct"]), 2)
        row = [r["Reporting_Airline"], int(r["total"]), otp_val, round(float(r.get("retraso_prom") or 0), 2)]
        if has_dep:
            row.append(round(float(r.get("dep_pct") or 0), 2))
        ws.append(row)
        ri = ws.max_row
        _color_otp(ws, ri, OTP_COL, otp_val)
        _stripe(ws, ri, len(headers))

    _autofit(ws)
    ws.freeze_panes = "A2"

    # Bar chart OTP por aerolínea
    if ws.max_row > 2:
        chart = BarChart()
        chart.type = "col"
        chart.title = "OTP por Aerolínea (%)"
        chart.y_axis.title = "OTP %"
        chart.y_axis.scaling.min = 0
        chart.y_axis.scaling.max = 100
        chart.shape = 4
        data = Reference(ws, min_col=OTP_COL, min_row=1, max_col=OTP_COL, max_row=ws.max_row)
        cats = Reference(ws, min_col=1, min_row=2, max_row=ws.max_row)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        chart.width = 26
        chart.height = 14
        anchor = f"A{ws.max_row + 2}"
        ws.add_chart(chart, anchor)


def _sheet_tendencia(wb, df) -> None:
    """Hoja: Tendencia mensual de vuelos, cancelaciones y OTP + line chart."""
    ws = wb.create_sheet("Tendencia Mensual")
    if "Month" not in df.columns or "pk_vuelo" not in df.columns:
        ws["A1"] = "Datos no disponibles"
        return

    grp = df.groupby("Month").agg(total=("pk_vuelo", "count")).reset_index()
    if "Cancelled" in df.columns:
        cm = df.groupby("Month")["Cancelled"].sum().reset_index(name="cancelados")
        grp = grp.merge(cm, on="Month", how="left")
        grp["cancelados"] = grp["cancelados"].fillna(0).astype(int)
    if "ArrDel15" in df.columns and "Cancelled" in df.columns:
        vop = df[df["Cancelled"] == 0].copy()
        vop["_at"] = (vop["ArrDel15"] == 0).astype(int)
        om = vop.groupby("Month").agg(op=("pk_vuelo", "count"), at=("_at", "sum")).reset_index()
        om["otp_pct"] = (om["at"] / om["op"].replace(0, 1) * 100).round(1)
        grp = grp.merge(om[["Month", "otp_pct"]], on="Month", how="left")
    if "ArrDelayMinutes" in df.columns and "Cancelled" in df.columns:
        vop2 = df[df["Cancelled"] == 0]
        dm = vop2.groupby("Month")["ArrDelayMinutes"].mean().round(1).reset_index(name="retraso_prom")
        grp = grp.merge(dm, on="Month", how="left")

    has_canc = "cancelados" in grp.columns
    has_otp = "otp_pct" in grp.columns
    has_delay = "retraso_prom" in grp.columns

    headers = ["Mes", "Total vuelos"]
    if has_canc:
        headers += ["Cancelados", "Tasa cancel. (%)"]
    if has_otp:
        headers.append("OTP (%)")
    if has_delay:
        headers.append("Retraso prom. arr. (min)")
    _hdr(ws, 1, headers)

    OTP_COL_IDX = (headers.index("OTP (%)") + 1) if has_otp else None
    for _, r in grp.sort_values("Month").iterrows():
        mn = int(r["Month"])
        row = [_MESES_LABEL[mn] if 1 <= mn <= 12 else str(mn), int(r["total"])]
        if has_canc:
            canc = int(r["cancelados"])
            row += [canc, round(canc / max(int(r["total"]), 1) * 100, 1)]
        if has_otp:
            row.append(round(float(r.get("otp_pct") or 0), 1))
        if has_delay:
            row.append(round(float(r.get("retraso_prom") or 0), 1))
        ws.append(row)
        ri = ws.max_row
        _stripe(ws, ri, len(headers))
        if OTP_COL_IDX and has_otp:
            _color_otp(ws, ri, OTP_COL_IDX, float(r.get("otp_pct") or 0))

    _autofit(ws)
    ws.freeze_panes = "A2"

    # Line chart OTP mensual
    if has_otp and ws.max_row > 2 and OTP_COL_IDX:
        chart = LineChart()
        chart.title = "OTP Mensual (%)"
        chart.y_axis.title = "OTP %"
        chart.y_axis.scaling.min = 50
        chart.y_axis.scaling.max = 100
        chart.smooth = True
        data = Reference(ws, min_col=OTP_COL_IDX, min_row=1, max_row=ws.max_row)
        cats = Reference(ws, min_col=1, min_row=2, max_row=ws.max_row)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        chart.width = 24
        chart.height = 13
        anchor_col = get_column_letter(len(headers) + 2)
        ws.add_chart(chart, f"{anchor_col}2")


def _sheet_causas_retraso(wb, filtros: dict | None) -> None:
    """Hoja: Minutos de retraso por causa y aerolínea (desde agg_causas_retraso_mes)."""
    ws = wb.create_sheet("Causas de Retraso")
    try:
        df = load_agg("agg_causas_retraso_mes", filtros)
    except Exception:
        ws["A1"] = "Datos no disponibles (ejecute el pipeline primero)"
        return

    if df.empty:
        ws["A1"] = "Sin datos para el período seleccionado"
        return

    # Totales por causa (para el resumen)
    ws["A1"] = "Distribución global de causas de retraso (minutos totales)"
    ws["A1"].font = Font(bold=True, size=11, color=_BLUE_DARK)
    ws.append([])

    resumen_headers = ["Causa", "Minutos totales", "% del total"]
    _hdr(ws, 3, resumen_headers)
    totales = []
    for col, label in _CAUSA_COLS:
        if col in df.columns:
            total = float(df[col].fillna(0).sum())
            totales.append((label, total))
    total_all = sum(v for _, v in totales) or 1
    for ri, (lbl, val) in enumerate(sorted(totales, key=lambda x: -x[1]), 4):
        pct = round(val / total_all * 100, 1)
        ws.append([lbl, round(val, 0), pct])
        _stripe(ws, ri, 3)

    # Bar chart causas
    data_row_start = 3
    data_row_end = 3 + len(totales)
    if data_row_end > data_row_start:
        chart = BarChart()
        chart.type = "bar"
        chart.title = "Minutos de Retraso por Causa"
        chart.x_axis.title = "Minutos"
        chart.shape = 4
        data_ref = Reference(ws, min_col=2, min_row=data_row_start, max_row=data_row_end)
        cats_ref = Reference(ws, min_col=1, min_row=data_row_start + 1, max_row=data_row_end)
        chart.add_data(data_ref, titles_from_data=True)
        chart.set_categories(cats_ref)
        chart.width = 24
        chart.height = 12
        ws.add_chart(chart, "E3")

    # Tabla por aerolínea
    row_start = data_row_end + 3
    ws.cell(row_start, 1, "Desglose por aerolínea").font = Font(bold=True, size=11, color=_BLUE_DARK)
    row_start += 1

    avail_cols = [(col, lbl) for col, lbl in _CAUSA_COLS if col in df.columns]
    al_headers = ["Aerolínea"] + [lbl for _, lbl in avail_cols] + ["Total retraso (min)"]
    _hdr(ws, row_start, al_headers)
    row_start += 1

    if "carrier" in df.columns:
        grp = df.groupby("carrier")[[col for col, _ in avail_cols]].sum().reset_index()
        grp["total"] = grp[[col for col, _ in avail_cols]].sum(axis=1)
        grp = grp.sort_values("total", ascending=False)
        for ri_offset, (_, r) in enumerate(grp.iterrows(), row_start):
            row = [r["carrier"]] + [round(float(r.get(col) or 0), 0) for col, _ in avail_cols]
            row.append(round(float(r["total"]), 0))
            ws.append(row)
            _stripe(ws, ri_offset, len(al_headers))

    _autofit(ws)


def _sheet_peores_rutas(wb, filtros: dict | None) -> None:
    """Hoja: Rutas con menor OTP (desde agg_rutas_eficiencia)."""
    ws = wb.create_sheet("Rutas — Peores OTP")
    try:
        df = load_agg("agg_rutas_eficiencia", filtros)
    except Exception:
        ws["A1"] = "Datos no disponibles (ejecute el pipeline primero)"
        return

    if df.empty or "otp_pct" not in df.columns:
        ws["A1"] = "Sin datos para el período seleccionado"
        return

    # Agrupar a nivel ruta (origin-dest) sumando carriers
    if "origin" in df.columns and "dest" in df.columns and "total_vuelos" in df.columns:
        grp = (
            df.groupby(["origin", "dest"])
            .agg(
                total_vuelos=("total_vuelos", "sum"),
                vuelos_at=("vuelos_a_tiempo", "sum") if "vuelos_a_tiempo" in df.columns else ("total_vuelos", "sum"),
                retraso_prom=("retraso_prom", "mean") if "retraso_prom" in df.columns else ("total_vuelos", "sum"),
            )
            .reset_index()
        )
        if "vuelos_a_tiempo" in df.columns:
            grp["otp_pct"] = (grp["vuelos_at"] / grp["total_vuelos"].replace(0, 1) * 100).round(1)
        else:
            grp["otp_pct"] = 0.0
        peores = grp[grp["total_vuelos"] >= 50].sort_values("otp_pct").head(30)
    else:
        peores = df.sort_values("otp_pct").head(30)

    headers = ["Origen", "Destino", "Ruta", "Vuelos", "OTP (%)", "Retraso prom. (min)"]
    _hdr(ws, 1, headers)
    OTP_COL = 5

    for _, r in peores.iterrows():
        orig = r.get("origin", "")
        dest = r.get("dest", "")
        otp = round(float(r.get("otp_pct") or 0), 1)
        ws.append(
            [
                orig,
                dest,
                f"{orig}-{dest}",
                int(r.get("total_vuelos") or 0),
                otp,
                round(float(r.get("retraso_prom") or 0), 1),
            ]
        )
        ri = ws.max_row
        _color_otp(ws, ri, OTP_COL, otp)
        _stripe(ws, ri, len(headers))

    _autofit(ws)
    ws.freeze_panes = "A2"

    # Bar chart peores rutas
    if ws.max_row > 2:
        chart = BarChart()
        chart.type = "bar"
        chart.title = "Rutas con Menor OTP (%)"
        chart.x_axis.title = "OTP %"
        chart.shape = 4
        data = Reference(ws, min_col=OTP_COL, min_row=1, max_row=ws.max_row)
        cats = Reference(ws, min_col=3, min_row=2, max_row=ws.max_row)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        chart.width = 26
        chart.height = 16
        ws.add_chart(chart, "H2")


def _sheet_dia_semana(wb, filtros: dict | None) -> None:
    """Hoja: OTP por día de semana (desde agg_otp_dia_semana)."""
    ws = wb.create_sheet("OTP por Día de Semana")
    try:
        df = load_agg("agg_otp_dia_semana")
    except Exception:
        ws["A1"] = "Datos no disponibles (ejecute el pipeline primero)"
        return

    if df.empty or "otp_pct" not in df.columns:
        ws["A1"] = "Sin datos"
        return

    df = df.sort_values("day_of_week")
    headers = ["Día de semana", "Total vuelos", "Vuelos a tiempo", "OTP (%)"]
    _hdr(ws, 1, headers)
    OTP_COL = 4

    for _, r in df.iterrows():
        dow = int(r["day_of_week"])
        otp = round(float(r["otp_pct"]), 1)
        ws.append(
            [
                _DOW_LABEL[dow] if 1 <= dow <= 7 else str(dow),
                int(r.get("total_vuelos") or 0),
                int(r.get("vuelos_a_tiempo") or 0),
                otp,
            ]
        )
        ri = ws.max_row
        _color_otp(ws, ri, OTP_COL, otp)
        _stripe(ws, ri, len(headers))

    _autofit(ws)

    # Bar chart OTP por día
    if ws.max_row > 2:
        chart = BarChart()
        chart.type = "col"
        chart.title = "OTP por Día de Semana (%)"
        chart.y_axis.title = "OTP %"
        chart.y_axis.scaling.min = 0
        chart.y_axis.scaling.max = 100
        chart.shape = 4
        data = Reference(ws, min_col=OTP_COL, min_row=1, max_row=ws.max_row)
        cats = Reference(ws, min_col=1, min_row=2, max_row=ws.max_row)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        chart.width = 22
        chart.height = 13
        ws.add_chart(chart, "F2")


def _sheet_cancelaciones(wb, df) -> None:
    """Hoja: Cancelaciones por causa FAA."""
    ws = wb.create_sheet("Cancelaciones FAA")
    if "CancellationCode" not in df.columns or "Cancelled" not in df.columns:
        ws["A1"] = "Datos no disponibles"
        return

    cancelados = df[df["Cancelled"] == 1]
    total_canc = max(len(cancelados), 1)
    grp = cancelados.groupby("CancellationCode").size().reset_index(name="count")
    headers = ["Código FAA", "Descripción", "Cancelaciones", "% del total"]
    _hdr(ws, 1, headers)

    for _, r in grp.sort_values("count", ascending=False).iterrows():
        code = str(r["CancellationCode"])
        pct = round(int(r["count"]) / total_canc * 100, 2)
        ws.append([code, _FAA_DESC.get(code, "Otro"), int(r["count"]), pct])
        _stripe(ws, ws.max_row, len(headers))

    _autofit(ws)

    # Pie chart cancelaciones
    if ws.max_row > 2:
        from openpyxl.chart import PieChart

        chart = PieChart()
        chart.title = "Cancelaciones por Causa FAA"
        data = Reference(ws, min_col=3, min_row=1, max_row=ws.max_row)
        cats = Reference(ws, min_col=2, min_row=2, max_row=ws.max_row)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        chart.width = 18
        chart.height = 13
        ws.add_chart(chart, "F2")


def _sheet_rutas_eficiencia(wb, df) -> None:
    """Hoja: Top rutas eficientes."""
    ws = wb.create_sheet("Rutas Eficientes")
    if "OriginCode" not in df.columns or "DestCode" not in df.columns:
        ws["A1"] = "Datos no disponibles"
        return

    vuelos_op = df[df["Cancelled"] == 0] if "Cancelled" in df.columns else df
    if "ActualElapsedTime" not in vuelos_op.columns or "CRSElapsedTime" not in vuelos_op.columns:
        ws["A1"] = "Columnas de tiempo no disponibles"
        return

    sub = vuelos_op[vuelos_op["CRSElapsedTime"] > 0].copy()
    sub["eficiencia"] = sub["ActualElapsedTime"] / sub["CRSElapsedTime"]
    grp = (
        sub.groupby(["OriginCode", "DestCode"])
        .agg(total=("pk_vuelo", "count"), ef=("eficiencia", "mean"))
        .reset_index()
    )
    grp = grp[grp["total"] >= 10].sort_values("ef").head(30)
    headers = ["Ruta", "Origen", "Destino", "Vuelos", "Índice eficiencia"]
    _hdr(ws, 1, headers)
    EF_COL = 5

    for _, r in grp.iterrows():
        ef = round(float(r["ef"]), 4)
        ws.append(
            [
                f"{r['OriginCode']}-{r['DestCode']}",
                r["OriginCode"],
                r["DestCode"],
                int(r["total"]),
                ef,
            ]
        )
        ri = ws.max_row
        _stripe(ws, ri, len(headers))
        cell = ws.cell(row=ri, column=EF_COL)
        if ef <= 1.05:
            cell.fill = PatternFill("solid", fgColor=_GREEN)
            cell.font = Font(bold=True, color="FFFFFF", size=10)
        elif ef <= 1.15:
            cell.fill = PatternFill("solid", fgColor=_AMBER)
            cell.font = Font(bold=True, color="FFFFFF", size=10)
        else:
            cell.fill = PatternFill("solid", fgColor=_RED)
            cell.font = Font(bold=True, color="FFFFFF", size=10)

    _autofit(ws)
    ws.freeze_panes = "A2"


def _sheet_resumen(wb, df, filtros: dict | None) -> None:
    """Hoja de resumen ejecutivo."""
    ws = wb.create_sheet("Resumen Ejecutivo", 0)
    title_fill = PatternFill("solid", fgColor=_BLUE_DARK)

    ws.merge_cells("A1:C1")
    ws["A1"] = "AeroTrack Analytics — Reporte Operacional"
    ws["A1"].font = Font(bold=True, size=14, color="FFFFFF")
    ws["A1"].fill = title_fill

    ws["A2"] = f"Generado: {datetime.now(_TZ).strftime('%d/%m/%Y %H:%M')}"
    ws["A2"].font = Font(size=10, italic=True, color="64748b")

    # Filtros
    ws.append([])
    row = 4
    ws.cell(row, 1, "FILTROS APLICADOS").font = Font(bold=True, size=11, color=_BLUE_DARK)
    row += 1
    for label, val in _filtros_display(filtros):
        ws.cell(row, 1, label).font = Font(bold=True, size=10)
        ws.cell(row, 2, val).font = Font(size=10)
        row += 1

    # Métricas globales
    row += 1
    ws.cell(row, 1, "MÉTRICAS GLOBALES").font = Font(bold=True, size=11, color=_BLUE_DARK)
    row += 1

    total_vuelos = len(df)
    cancelados = int(df["Cancelled"].sum()) if "Cancelled" in df.columns else 0
    desviados = int(df["Diverted"].sum()) if "Diverted" in df.columns else 0
    otp_global = 0.0
    dep_delay_pct = 0.0
    if "ArrDel15" in df.columns and "Cancelled" in df.columns:
        vop = df[df["Cancelled"] == 0]
        otp_global = round(float((vop["ArrDel15"] == 0).sum() / max(len(vop), 1) * 100), 2)
    if "DepDel15" in df.columns and "Cancelled" in df.columns:
        vop2 = df[df["Cancelled"] == 0]
        dep_delay_pct = round(float((vop2["DepDel15"] == 1).sum() / max(len(vop2), 1) * 100), 2)
    retraso_arr = 0.0
    if "ArrDelayMinutes" in df.columns and "Cancelled" in df.columns:
        vop3 = df[df["Cancelled"] == 0]
        retraso_arr = round(float(vop3["ArrDelayMinutes"].mean() or 0), 2)

    metricas = [
        ("Total vuelos (incl. cancelados)", total_vuelos),
        ("Vuelos operados", total_vuelos - cancelados),
        ("Cancelados", cancelados),
        ("Tasa cancelación (%)", round(cancelados / max(total_vuelos, 1) * 100, 2)),
        ("Desviados", desviados),
        ("OTP global — arr. (%)", otp_global),
        ("% vuelos con demora salida ≥15 min", dep_delay_pct),
        ("Retraso promedio llegada (min)", retraso_arr),
    ]
    for label, val in metricas:
        ws.cell(row, 1, label).font = Font(bold=True, size=10)
        c = ws.cell(row, 2, val)
        c.font = Font(size=10)
        c.alignment = Alignment(horizontal="right")
        row += 1

    # Nota de hojas
    row += 1
    ws.cell(row, 1, "CONTENIDO DEL REPORTE").font = Font(bold=True, size=11, color=_BLUE_DARK)
    row += 1
    for hoja, desc in [
        ("Puntualidad OTP", "OTP por aerolínea + gráfico de columnas"),
        ("Tendencia Mensual", "Vuelos, cancelaciones y OTP mes a mes + gráfico"),
        ("Causas de Retraso", "Minutos de retraso por tipo de causa + gráfico"),
        ("Rutas — Peores OTP", "Rutas con menor OTP (problema operacional) + gráfico"),
        ("OTP por Día de Semana", "Patrón semanal de puntualidad + gráfico"),
        ("Cancelaciones FAA", "Distribución por código FAA + gráfico circular"),
        ("Rutas Eficientes", "Top rutas por eficiencia de tiempo real vs programado"),
    ]:
        ws.cell(row, 1, hoja).font = Font(bold=True, size=10)
        ws.cell(row, 2, desc).font = Font(size=10, color="334155")
        row += 1

    ws.column_dimensions["A"].width = 38
    ws.column_dimensions["B"].width = 48


# ── Public API ────────────────────────────────────────────────────────────────


def generar_excel(filtros: dict | None = None) -> bytes:
    """Genera .xlsx con 8 hojas, datos y gráficos embebidos. Retorna bytes."""
    df = load_enriched_fact(filtros)

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # hojas de datos analíticos (desde fact enriquecido)
    _sheet_otp(wb, df)
    _sheet_tendencia(wb, df)
    _sheet_cancelaciones(wb, df)
    _sheet_rutas_eficiencia(wb, df)

    # hojas desde tablas de agregación (más datos, sin cargar el fact completo)
    _sheet_causas_retraso(wb, filtros)
    _sheet_peores_rutas(wb, filtros)
    _sheet_dia_semana(wb, filtros)

    # hoja resumen al frente
    _sheet_resumen(wb, df, filtros)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()
